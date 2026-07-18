"""예산관리 엑셀 자동생성 스크립트.

항목명/예산액/집행액을 받아 잔액·집행률이 엑셀 수식으로 들어간 xlsx를 만든다.
집행률이 80% 이상이면 노란색, 100% 이상이면 빨간색으로 행 전체가 칠해진다.

값을 파이썬에서 계산해 박지 않고 수식과 조건부 서식으로 넣기 때문에,
엑셀에서 집행액만 고쳐도 잔액·집행률·경고색이 자동으로 따라온다.

사용법:
    python budget_excel.py          # 샘플 데이터로 sample_budget.xlsx 생성

    from budget_excel import create_budget_file
    create_budget_file([("사무용품비", 1_000_000, 450_000)], "예산.xlsx")
"""

from numbers import Real
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

HEADERS = ("항목명", "예산액", "집행액", "잔액", "집행률")

# 경고 임계값. ">=" 로 비교하므로 정확히 80.0%인 항목도 노랑에 포함된다.
WARNING_THRESHOLD = 0.8
DANGER_THRESHOLD = 1.0

WARNING_COLOR = "FFEB9C"  # 노랑
DANGER_COLOR = "FFC7CE"  # 빨강
HEADER_COLOR = "D9D9D9"  # 회색

MONEY_FORMAT = "#,##0"
RATE_FORMAT = "0.0%"

COLUMN_WIDTHS = {"A": 24, "B": 14, "C": 14, "D": 14, "E": 10}

HEADER_ROW = 1
FIRST_DATA_ROW = 2

# 경계값을 모두 포함한 샘플. 파일을 열면 규칙이 눈으로 검증된다.
SAMPLE_ITEMS = [
    ("사무용품비", 1_000_000, 450_000),  # 45.0%  색 없음
    ("교육훈련비", 2_000_000, 1_600_000),  # 80.0%  노랑 (하단 경계)
    ("출장여비", 3_000_000, 2_760_000),  # 92.0%  노랑
    ("회의비", 500_000, 500_000),  # 100.0% 빨강 (경계)
    ("소모품비", 800_000, 920_000),  # 115.0% 빨강 (초과)
    ("예비비", 0, 0),  # 0.0%   색 없음 (0으로 나누기 방어)
]


def _validate(items):
    """입력을 검증하고 (항목명, 예산액, 집행액) 리스트로 정규화한다."""
    if not items:
        # 데이터 행이 없으면 조건부 서식 범위와 SUBTOTAL 범위를 만들 수 없다.
        raise ValueError("예산 항목이 최소 하나는 있어야 합니다.")

    validated = []
    for index, item in enumerate(items, start=1):
        try:
            name, budget, executed = item
        except (TypeError, ValueError):
            raise ValueError(
                f"{index}번째 항목은 (항목명, 예산액, 집행액) 3개 값이어야 합니다: {item!r}"
            ) from None

        if not isinstance(name, str) or not name.strip():
            raise ValueError(f"{index}번째 항목의 항목명이 비어 있습니다: {name!r}")

        for label, amount in (("예산액", budget), ("집행액", executed)):
            # bool은 Real의 하위 타입이라 별도로 걸러낸다.
            if isinstance(amount, bool) or not isinstance(amount, Real):
                raise ValueError(
                    f"{name!r}의 {label}이 숫자가 아닙니다: {amount!r}"
                )
            if amount < 0:
                raise ValueError(f"{name!r}의 {label}이 음수입니다: {amount!r}")

        validated.append((name.strip(), budget, executed))

    return validated


def _write_header(sheet):
    fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    for column, title in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=column, value=title)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_row_formulas(sheet, row):
    """잔액과 집행률을 수식으로 넣는다. 두 곳(데이터 행, 합계 행)에서 쓴다."""
    balance = sheet.cell(row=row, column=4, value=f"=B{row}-C{row}")
    balance.number_format = MONEY_FORMAT

    # IF로 예산액 0을 방어하지 않으면 미배정 항목에서 #DIV/0! 이 뜬다.
    rate = sheet.cell(row=row, column=5, value=f"=IF(B{row}=0,0,C{row}/B{row})")
    rate.number_format = RATE_FORMAT


def _write_data(sheet, items):
    for offset, (name, budget, executed) in enumerate(items):
        row = FIRST_DATA_ROW + offset
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=budget).number_format = MONEY_FORMAT
        sheet.cell(row=row, column=3, value=executed).number_format = MONEY_FORMAT
        _write_row_formulas(sheet, row)


def _write_total(sheet, last_data_row):
    row = last_data_row + 1
    cell = sheet.cell(row=row, column=1, value="합계")
    cell.font = Font(bold=True)

    for column in (2, 3):
        letter = get_column_letter(column)
        # SUBTOTAL(109, ...)는 자동 필터로 숨긴 행을 제외한다.
        # 사용자가 일부 항목만 필터링해도 합계가 의미를 유지한다.
        total = sheet.cell(
            row=row,
            column=column,
            value=f"=SUBTOTAL(109,{letter}{FIRST_DATA_ROW}:{letter}{last_data_row})",
        )
        total.number_format = MONEY_FORMAT
        total.font = Font(bold=True)

    _write_row_formulas(sheet, row)
    for column in (4, 5):
        sheet.cell(row=row, column=column).font = Font(bold=True)

    top_border = Border(top=Side(style="thin"))
    for column in range(1, len(HEADERS) + 1):
        sheet.cell(row=row, column=column).border = top_border


def _apply_conditional_formatting(sheet, last_data_row):
    """집행률 기준 경고색을 규칙으로 등록한다.

    합계 행은 범위에서 제외한다. 전체 집행률이 80%를 넘었다고 합계 줄까지
    칠하면 어떤 개별 항목이 문제인지 읽기 어려워진다.
    """
    cell_range = f"A{FIRST_DATA_ROW}:E{last_data_row}"

    # $E2 — 열은 고정, 행은 상대. 이래야 한 규칙이 각 행에 자기 행 기준으로 적용된다.
    # 100% 이상인 행은 두 조건을 모두 만족하므로, 빨강을 먼저 등록해
    # 더 높은 우선순위를 갖게 한다.
    for threshold, color in (
        (DANGER_THRESHOLD, DANGER_COLOR),
        (WARNING_THRESHOLD, WARNING_COLOR),
    ):
        formula = f"$E{FIRST_DATA_ROW}>={threshold:g}"
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet.conditional_formatting.add(cell_range, FormulaRule(formula=[formula], fill=fill))


def create_budget_file(items, output_path):
    """예산관리 xlsx를 생성하고 저장된 경로를 반환한다.

    items: (항목명, 예산액, 집행액) 시퀀스의 리스트
    output_path: 저장 경로 (str 또는 Path)
    """
    validated = _validate(items)
    last_data_row = FIRST_DATA_ROW + len(validated) - 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "예산관리"

    _write_header(sheet)
    _write_data(sheet, validated)
    _write_total(sheet, last_data_row)
    _apply_conditional_formatting(sheet, last_data_row)

    for letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:E{last_data_row}"

    path = Path(output_path)
    workbook.save(path)
    return path


def main():
    path = create_budget_file(SAMPLE_ITEMS, Path(__file__).parent / "sample_budget.xlsx")
    print(f"생성 완료: {path}")
    # em 대시 같은 문자는 윈도우 콘솔(cp949)에서 UnicodeEncodeError를 낸다.
    print(f"항목 {len(SAMPLE_ITEMS)}개 / 집행률 80% 이상 노랑, 100% 이상 빨강")


if __name__ == "__main__":
    main()
