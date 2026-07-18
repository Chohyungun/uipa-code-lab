"""budget_excel 모듈 검증 테스트.

생성된 xlsx를 openpyxl로 다시 읽어 수식과 서식이 올바르게 기록됐는지 확인한다.
조건부 서식의 실제 렌더링 색과 수식의 계산 결과는 엑셀이 파일을 열 때
계산하므로 여기서는 검증할 수 없다 (docs/.../design.md의 '검증의 한계' 참고).
"""

import pytest
from openpyxl import load_workbook

from budget_excel import HEADERS, SAMPLE_ITEMS, create_budget_file

ITEMS = [
    ("사무용품비", 1_000_000, 450_000),
    ("교육훈련비", 2_000_000, 1_600_000),
    ("회의비", 500_000, 500_000),
]


@pytest.fixture
def sheet(tmp_path):
    """ITEMS로 파일을 만들고 되읽은 워크시트."""
    path = create_budget_file(ITEMS, tmp_path / "budget.xlsx")
    return load_workbook(path).active


def normalize_color(rgb):
    """openpyxl은 색을 '00FFC7CE'처럼 알파 접두사를 붙여 되읽는다."""
    return rgb[-6:]


# --- 기본 구조 ---


def test_헤더가_지정된_순서로_기록된다(sheet):
    actual = [sheet.cell(row=1, column=i).value for i in range(1, 6)]
    assert actual == list(HEADERS)


def test_항목명과_금액이_그대로_기록된다(sheet):
    assert sheet["A2"].value == "사무용품비"
    assert sheet["B2"].value == 1_000_000
    assert sheet["C2"].value == 450_000
    assert sheet["A4"].value == "회의비"


def test_틀고정이_헤더_아래에_걸린다(sheet):
    assert sheet.freeze_panes == "A2"


# --- 수식 ---
# 계산된 숫자가 아니라 수식이 들어가야 엑셀에서 집행액을 고쳤을 때 따라온다.


def test_잔액은_계산된_값이_아니라_수식이다(sheet):
    assert sheet["D2"].value == "=B2-C2"
    assert sheet["D4"].value == "=B4-C4"


def test_집행률_수식은_예산액_0을_방어한다(sheet):
    assert sheet["E2"].value == "=IF(B2=0,0,C2/B2)"


def test_합계행은_SUBTOTAL로_집계한다(sheet):
    total_row = 2 + len(ITEMS)  # 데이터 3행 다음
    assert sheet.cell(row=total_row, column=1).value == "합계"
    assert sheet.cell(row=total_row, column=2).value == "=SUBTOTAL(109,B2:B4)"
    assert sheet.cell(row=total_row, column=3).value == "=SUBTOTAL(109,C2:C4)"


# --- 표시 형식 ---


def test_금액열은_천단위_구분기호를_쓴다(sheet):
    assert sheet["B2"].number_format == "#,##0"
    assert sheet["C2"].number_format == "#,##0"
    assert sheet["D2"].number_format == "#,##0"


def test_집행률열은_백분율로_표시된다(sheet):
    assert sheet["E2"].number_format == "0.0%"


# --- 조건부 서식 ---


def get_rules(sheet):
    """(범위문자열, 우선순위순 규칙목록) 반환."""
    ranges = list(sheet.conditional_formatting)
    assert len(ranges) == 1, "조건부 서식은 한 범위에만 적용되어야 한다"
    cf = ranges[0]
    return str(cf.sqref), sorted(cf.rules, key=lambda r: r.priority)


def test_조건부서식은_데이터행에만_걸리고_합계행은_제외한다(sheet):
    sqref, _ = get_rules(sheet)
    assert sqref == "A2:E4"  # 합계행(5행) 미포함


def test_빨강규칙이_노랑보다_우선한다(sheet):
    """100% 이상인 행은 두 조건을 모두 만족하므로 순서가 결과를 결정한다."""
    _, rules = get_rules(sheet)
    assert len(rules) == 2
    assert rules[0].formula == ["$E2>=1"]
    assert rules[1].formula == ["$E2>=0.8"]


def test_규칙의_색이_빨강과_노랑이다(sheet):
    _, rules = get_rules(sheet)
    assert normalize_color(rules[0].dxf.fill.bgColor.rgb) == "FFC7CE"
    assert normalize_color(rules[1].dxf.fill.bgColor.rgb) == "FFEB9C"


def test_규칙은_열고정_행상대_참조를_쓴다(sheet):
    """$E2 — 열 고정이 빠지면 규칙이 각 행에 잘못 적용된다."""
    _, rules = get_rules(sheet)
    for rule in rules:
        assert rule.formula[0].startswith("$E2")


# --- 입력 검증 ---


@pytest.mark.parametrize(
    "bad_items, reason",
    [
        ([], "빈 목록"),
        ([("", 100, 50)], "빈 항목명"),
        ([("   ", 100, 50)], "공백뿐인 항목명"),
        ([("항목", -100, 50)], "음수 예산액"),
        ([("항목", 100, -50)], "음수 집행액"),
        ([("항목", "백만원", 50)], "숫자가 아닌 예산액"),
        ([("항목", 100, None)], "숫자가 아닌 집행액"),
    ],
)
def test_잘못된_입력은_ValueError를_낸다(tmp_path, bad_items, reason):
    with pytest.raises(ValueError):
        create_budget_file(bad_items, tmp_path / "bad.xlsx")


def test_예산액_0은_허용된다(tmp_path):
    """미배정 항목은 실제로 존재하고, 집행률 수식이 IF로 방어한다."""
    path = create_budget_file([("예비비", 0, 0)], tmp_path / "zero.xlsx")
    sheet = load_workbook(path).active
    assert sheet["B2"].value == 0
    assert sheet["E2"].value == "=IF(B2=0,0,C2/B2)"


# --- 샘플 데이터 ---


def test_샘플데이터는_경계값을_모두_포함한다(tmp_path):
    """파일을 열었을 때 규칙이 눈으로 검증되도록 경계값이 들어가야 한다."""
    rates = [executed / budget for _, budget, executed in SAMPLE_ITEMS if budget]
    assert any(r < 0.8 for r in rates), "색 없는 정상 항목"
    assert any(r == 0.8 for r in rates), "노랑 하단 경계"
    assert any(0.8 < r < 1.0 for r in rates), "노랑"
    assert any(r == 1.0 for r in rates), "빨강 경계"
    assert any(r > 1.0 for r in rates), "빨강 초과"
    assert any(budget == 0 for _, budget, _ in SAMPLE_ITEMS), "예산액 0"


def test_샘플데이터로_파일이_생성된다(tmp_path):
    path = create_budget_file(SAMPLE_ITEMS, tmp_path / "sample.xlsx")
    assert path.exists()
    sheet = load_workbook(path).active
    assert sheet.max_row == len(SAMPLE_ITEMS) + 2  # 헤더 + 데이터 + 합계
